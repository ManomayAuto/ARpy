from openpyxl import *
import pyodbc

from datetime import datetime
import os
import xlrd
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Alignment
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=INBROKER4;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))




curdate = (datetime.today()).strftime('%d')
print(curdate)
if curdate == '1':
    print("true")

    wb = Workbook()
    ws = wb.active
    ws.freeze_panes = ws['A2']
    ws.cell(1, 1).value = ('User')
    ws['A1'].font = Font(bold=True)
    ws.cell(1, 1).border = thin_border
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 2).value = ('User  Type')
    ws['B1'].font = Font(bold=True)
    ws.cell(1, 2).border = thin_border
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 3).value = ('Date')
    ws['C1'].font = Font(bold=True)
    ws.cell(1, 3).border = thin_border
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 4).value = ('Follow ups Completed')
    ws['D1'].font = Font(bold=True)
    ws.cell(1, 4).border = thin_border
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 5).value = ('Reductions Completed')
    ws['E1'].font = Font(bold=True)
    ws.cell(1, 5).border = thin_border
    ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    row = 2

    csr = conn.cursor()
    print(datetime.today())
    dt = datetime.today()
    print(dt)
    dt1 = (dt).strftime('%Y-%m-%d')
    print(dt1)

    csr.execute(
        " Select um.name,max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like '%DF,DR%' union select ae.accexe,count(lastfollowup) as f,count(DateCompleted) as r from ae where cast(lastfollowup as date) = ? or cast(DateCompleted as date) = ? group by ae.accexe) a on um.name=a.name and cast(createddate as date)<= ? group by um.name ",
        dt1, dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    AErecords = len(rcrds)
    print(AErecords)

    aeName = 'AE'

    col = 1
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = aeName
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[1]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[2]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        print(r[0])
        print(r[1])

    # NONAE DATA - Followups

    csr = conn.cursor()
    csr.execute(
        " Select um.name,'Account',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DF' union select nonae.username,count(followups) as f,0 as r from NONAE where cast(DateCompleted as date) = ? and followups=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",
        dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    NONAEfrecords = len(rcrds) + AErecords
    # if AErecords == 0 :
    #     row = 1
    # else:
    #     row = row
    # col = 0
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[2]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[3]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    # NONAE DATA - Reductions

    csr = conn.cursor()
    csr.execute(
        "Select um.name,'CS',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DR' union select nonae.username,0 as f,count(ReductionCheckBox) as r from NONAE where cast(DateCompleted as date) =? and ReductionCheckBox=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",
        dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    NONAERrecords = len(rcrds) + NONAEfrecords
    # if NONAEfrecords == 0 :
    #     row = 1
    # else:
    #     row = row
    # col = 0
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[2]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[3]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width
    repdate = datetime.today().strftime('%b%Y')
    wb.save('C:/Users/administrator.JSJHO/Desktop/Consultant Files/Mantra/Consolidatedactivity/ConsolidatedActivityReport_'+dt+'.xlsx')

else:
    dt = (datetime.today()).strftime('%b%Y')
    print(dt)
    t = os.path.exists('C:/Users/administrator.JSJHO/Desktop/Consultant Files/Mantra/Consolidatedactivity/ConsolidatedActivityReport_'+dt+'.xlsx')
    print(t)

    if t:
        book = xlrd.open_workbook('C:/Users/administrator.JSJHO/Desktop/Consultant Files/Mantra/Consolidatedactivity/ConsolidatedActivityReport_'+dt+'.xlsx')
        sheet = book.sheet_by_index(0)
        nr = sheet.nrows
        print(nr)
        wb = load_workbook('C:/Users/administrator.JSJHO/Desktop/Consultant Files/Mantra/Consolidatedactivity/ConsolidatedActivityReport_'+dt+'.xlsx')
        ws = wb["Sheet"]
        row = nr + 1

    else:
        wb = Workbook()
        ws = wb.active
        ws.freeze_panes = ws['A2']
        ws.cell(1, 1).value = ('User')
        ws['A1'].font = Font(bold=True)
        ws.cell(1, 1).border = thin_border
        ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 2).value = ('User  Type')
        ws['B1'].font = Font(bold=True)
        ws.cell(1, 2).border = thin_border
        ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
        ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 3).value = ('Date')
        ws['C1'].font = Font(bold=True)
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 4).value = ('Follow ups Completed')
        ws['D1'].font = Font(bold=True)
        ws.cell(1, 4).border = thin_border
        ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        ws.cell(1, 5).value = ('Reductions Completed')
        ws['E1'].font = Font(bold=True)
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
        row = 2

    csr = conn.cursor()


    dt1 = datetime.today().strftime('%Y-%m-%d')
    print(dt1)

    csr.execute(
        " Select um.name,max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like '%DF,DR%' union select ae.accexe,count(lastfollowup) as f,count(DateCompleted) as r from ae where cast(lastfollowup as date) = ? or cast(DateCompleted as date) = ? group by ae.accexe) a on um.name=a.name and cast(createddate as date)<= ? group by um.name ",
        dt1, dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    AErecords = len(rcrds)
    print(AErecords)

    aeName = 'AE'

    col = 1
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = aeName
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[1]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[2]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1
        print(r[0])
        print(r[1])

    # NONAE DATA - Followups

    csr = conn.cursor()
    csr.execute(
        " Select um.name,'Account',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DF' union select nonae.username,count(followups) as f,0 as r from NONAE where cast(DateCompleted as date) = ? and followups=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",
        dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    NONAEfrecords = len(rcrds) + AErecords
    # if AErecords == 0 :
    #     row = 1
    # else:
    #     row = row
    # col = 0
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[2]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[3]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    # NONAE DATA - Reductions

    csr = conn.cursor()
    csr.execute(
        "Select um.name,'CS',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DR' union select nonae.username,0 as f,count(ReductionCheckBox) as r from NONAE where cast(DateCompleted as date) =? and ReductionCheckBox=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",
        dt1, dt1)
    rcrds = csr.fetchall()
    print(rcrds)
    NONAERrecords = len(rcrds) + NONAEfrecords
    # if NONAEfrecords == 0 :
    #     row = 1
    # else:
    #     row = row
    # col = 0
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = dt1
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[2]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[3]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width
    wb.save('C:/Users/administrator.JSJHO/Desktop/Consultant Files/Mantra/Consolidatedactivity/ConsolidatedActivityReport_'+dt+'.xlsx')