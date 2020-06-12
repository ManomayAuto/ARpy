import gc
import random
import smtplib

import json

from openpyxl import *

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import Flask, render_template, url_for, request, redirect, Response, session, flash, send_from_directory, \
    send_file, abort, jsonify, logging, make_response
from flask_cors import CORS, cross_origin

import jwt

from functools import wraps
from motor import getpremiummotor
from private import getpremiumprivate
from encryption import DataEncryption

import pyodbc
import openpyxl
import xlsxwriter
from datetime import datetime, timedelta
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Alignment

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')

# cred = credentials.Certificate("secret/portals-2edf2-firebase-adminsdk-aljih-9a4318b0c8.json")
# firebase_admin.initialize_app(cred)
# db1 = firestore.client()

app = Flask(__name__)
CORS(app)
permissions = ["QI", "QRU", "QRM", "QRP", "DF", "DR", "ARRP"]
de = DataEncryption()

def login_required(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'logged_in' in session:
            return f(*args, **kwargs)
        else:
            flash("You need to login first")
            return redirect(url_for('login'))

    return wrap

@app.route('/calculation', methods=['POST', 'OPTIONS'])
@cross_origin(
    origin=['*'], allow_headers=['Content-Type', 'Authorization', 'User'])
def cal():
    data = request.json
    print(data)
    product = data['prod']
    if product is not None and product != '':
        if product == 'Private':
            annualgpwp, annualgpwpd, annualnet, taxamt, nc = getpremiumprivate(data['cc'], data['vehiclevalue'], data['softd'], data['vehicletype'], data['manloadp'], data['claimfre'], data['manualdisc'], data['fleet'], data['promotion'], data['ct'])
            return jsonify({
                'annualgpwp': annualgpwp,
                'annualgpwpd': annualgpwpd,
                'annualnet': annualnet,
                'taxamt': taxamt,
                'nc': nc,
            })
        elif product == 'Motor':
            annualgpwp, annualgpwpd, annualnet, taxamt, nc = getpremiummotor(data['cc'], data['vehiclevalue'], data['softd'], data['vehicletype'], data['manloadp'], data['claimfre'], data['manualdisc'], data['fleet'], data['promotion'], data['ct'])
            return jsonify({
                'annualgpwp': annualgpwp,
                'annualgpwpd': annualgpwpd,
                'annualnet': annualnet,
                'taxamt': taxamt,
                'nc': nc,
            })
        else:
            return jsonify({"result": "please send a product"}),400
    else:
        return jsonify({"result": "please send a product"}),422
    return jsonify({"result": "success"}),200

@app.route('/checktoken', methods=['POST', 'OPTIONS'])
# @cross_origin(
#     origin=['*'], allow_headers=['Content-Type', 'Authorization', 'User'])
def checktoken():
    if request.method == "POST":
        result = False
        print(request.headers)
        try:
            token = ""
            request_keys = request.headers.keys()
            if "Authorization" in request_keys and "Ipaddress" in request_keys and "User" in request_keys:
                token = request.headers["Authorization"]
                print(token)
            print(request.get_data())
            data = json.loads(str(request.data, encoding='utf-8'))
            decoded = jwt.decode(token, key='secret')
            print('-------------------------')
            print(decoded)
            print(decoded["user"])
            if decoded["user"] == request.headers["User"] and decoded["ip"] == request.headers["Ipaddress"]:
                result = True
        except jwt.DecodeError:
            print("decode error")
            result = False
        except jwt.ExpiredSignatureError:
            print("sign")
            result = False
        except KeyError:
            print("key error")
            print(result)
        if result:
            return jsonify({"result": result}), 200
        else:
            return jsonify({"result": result}), 401

# this is for admin portal
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == "GET":
        return render_template("login.html")
    elif request.method == "POST":
        email = request.form["email"]
        password = request.form["password"]
        encryptpas = de.encrypt(password)
        print(email)
        print(password)
        cursor = conn.cursor()

        cursor.execute( "select Permissions,temppass,Name,Email from Users_man where Email = '%s' and curPassword = '%s'" % (email,encryptpas))

        temp_pass = 0
        permissions = None
        name = None

        for row in cursor:
            permissions = row[0]
            temp_pass = row[1]
            name = row[2]
            email = row[3]
        if temp_pass == 1:
            return redirect(url_for("resetpass"))
        if permissions is not None and name is not None and "Admin" in permissions:
            session["logged_in"] = True
            session["email"] = email
            return redirect(url_for('usermanagement'))
        else:
            flash("You have no Authorization")
            return redirect(url_for('login'))

#t this is for angular
@app.route('/login', methods=['POST'])
@cross_origin(origins="*", headers="*")
def login_form():
    print(request.data)
    data = json.loads(request.data, encoding='utf-8')
    # email = request.form["email"]
    # password = request.form["password"]
    email = data["email"]
    password = data["password"]
    print(email)
    print(password)
    encryptpas = de.encrypt(password)
    print(de.encrypt(password))
    cursor = conn.cursor()
    print("In Angular")
    print(email)
    print(password)
    cursor.execute("select Permissions,temppass,Name,Email,Branch from Users_man where Email = '%s' and curPassword = '%s'" % (email,encryptpas))

    temp_pass = True
    permissions = None
    name = None


    for row in cursor:
        permissions = row[0]
        temp_pass = row[1]
        name = row[2]
        email = row[3]
        branch = row[4]

    print(name)
    print(permissions)
    tp = 0
    print(temp_pass)
    if temp_pass:
        tp=1
    else:
        tp=0


    cursor.execute("Select * from users_man where Email = '%s' and deleted = 1"% email)
    if len(list(cursor)) > 0:
        print("In deleted if")
        flash('Your account has been disable. Please contact your administrator')
        #return redirect(url_for('login_form'))
        return jsonify({"message": "Your account has been disable. Please contact your administrator"}),401

    # elif temp_pass ==1:
    #     print("Logging in for the first time")
    #     return jsonify({"Logging in for the first time"}), 200

    elif permissions is not None and name is not None:

        current_time = datetime.now().timestamp()
        #current_time = datetime.now()
        print("current time -----------------------------------------------")
        print(current_time)
        #exp = (current_time + timedelta(hours=3)).timestamp()
        exp = datetime.utcnow() + timedelta(hours=3, minutes=0, seconds=0)
        print("Expired time ---------------------------------------------")
        print(exp)
        token = jwt.encode({"exp": exp, "iat": current_time, "permissions": permissions,"Name": name, "Email": email,"Branch": branch}, algorithm='HS256', key='secret')
        print(token)
        return jsonify({"permissions": permissions, "Email": email, "Name": name,"Branch": branch, "temppass": tp, "Token": token.decode("utf-8")}), 200
    else:

        print("Not authorized")
        return jsonify({"message": "Incorrect or invalid credentials"}),401



@app.route('/Usermanage', methods=['GET', 'POST'])
@login_required
def usermanagement():
    cursor = conn.cursor()
    cursor.execute(
        "select Name,Permissions,Email,temppass,lastupdated,branch,deleted from Users_man where Email <> '%s'" % (
            session.get('email')))
    names = []
    permissions = []
    emails = []
    branches = []
    deleted = []
    c = conn.cursor()
    for row in cursor:
        names.append(row[0])
        permissions.append(row[1])
        emails.append(row[2])
        branches.append(row[5])
        deleted.append(row[6])

    count = c.execute("select count(*) from Users_man where Email <> '%s'" % (
        session.get('email')))
    print(count)
    print(deleted)
    count = [int(i[0]) for i in c.fetchall()]
    print(count[0])
    cp = count[0]
    j = 0
    for i in deleted:
        if(i == 0):
            deleted[j] = "Active"
        else:
            deleted[j] = "Inactive"
        j = j+1
    print(deleted)
    return render_template("Usermanage.html", u=names, u1=permissions, u2=emails, u3=branches, dl=deleted, cp=cp, level=session.get('level'))


@app.route("/createuser", methods=['POST'])
@login_required
def add_user():
    try:
        if request.method == "POST":
            email = request.form.get("email")
            name = request.form.get("Name")
            password = request.form.get("Password")
            branch = request.form.get("branch")
            permissions_to_db = ""
            for permission in permissions:
                if request.form.get(permission):
                    permissions_to_db += permission + ","
            permissions_to_db = permissions_to_db[:-1]
            print(email)
            print(name)
            print(password)
            print(permissions_to_db)
            encryptpas = de.encrypt(password)
            print("The encrypted password is  -----------------------------------------")
            print(encryptpas)
            lastupdated = datetime.today().strftime('%Y-%m-%d')
            print("Afterecnrypt")
            createddate = datetime.today().strftime('%Y-%m-%d')
            cursor = conn.cursor()
            cursor.execute("INSERT INTO Users_man (Name,Permissions,Email,curPassword,Branch,Createdby,temppass,createddate,lastupdated,deleted) VALUES('%s','%s',"
                       "'%s','%s','%s','%s',%s,'%s','%s',%s)" % (name, permissions_to_db, email, encryptpas,branch, session.get("email"), 1,createddate,lastupdated,0))
            conn.commit()
            db1 = conn.cursor()
            db1.execute("Select name from users_man where Email = '%s'"% email)
            print("AfterDb")
            for row in db1:
                name = row[0]
                print(name)
                smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', port=465)
                smtpObj.login('praneeth.amudala@manomay.biz', "manomay72020")
                msg = MIMEMultipart()
                msg['subject'] = "JSJ Account Security Details"
                msg['from'] = 'praneeth.amudala@manomay.biz'
                msg['to'] = email
                msgtext1 = MIMEText(
                    '<p>Hi %s,</p>'
                    '<p>Your account has been registered on Integrated System for Process Automation.</p>'
                    '<p>Please use the following credentials to log in for the first-time</p>'
                    '<p><strong>Username : %s</strong></p>'
                    '<p>Password : %s</p>'
                    '<p style="color:Red;">Kindly reset your password on the first login.</p>'
                    '<p>Regards</p>'
                    '<p><strong>J.S.JOHNSON</strong></p>'%(name,email,password), 'html')
                msg.attach(msgtext1)
                smtpObj.sendmail('praneeth.amudala@manomay.biz', email, msg.as_string())
            print("Emailsend")
            flash("User created Successfully")
    except Exception as e:
        write_to_file(e)
        flash("Can't create the user")
        conn.rollback()
    finally:
        return redirect(url_for('usermanagement'))


@app.route('/updateuser', methods=['GET', 'POST'])
@login_required
def updateuser():
    email = request.form.get("Email")
    name = request.form.get("Name")
    status = request.form.get("colorRadio")
    branch = request.form.get("branch")
    permissions_to_db = ""
    for permission in permissions:
        if request.form.get(permission):
            permissions_to_db += permission + ","
    permissions_to_db = permissions_to_db[:-1]
    print(email)
    print(name)
    print(permissions_to_db)
    print(status)
    print(branch)

    if permissions_to_db == '' and status =='Inactive' :
        print("Hello")
        cursor = conn.cursor()
        sts = 1
        cursor.execute("Update users_man set deleted = '%s',lastupdated = '%s'  where Email='%s'" % (sts,datetime.today() ,email))
        conn.commit()
        flash("User has been inactivated")

        return redirect(url_for('usermanagement'))

    if permissions_to_db == '' and status =='Active' :
        print("Hello")

        return """<script>alert("You have not selected any permissions for the user. If you wish to remove access, please make the user inactive");window.location.href="/Usermanage"</script>"""

    else:
        cursor = conn.cursor()
        sts = 0
        if status == 'Inactive' :
            sts = 1
            cursor.execute("Update users_man set deleted = '%s',lastupdated = '%s' where Email='%s'" %(sts,datetime.today(),email))
            conn.commit()
            flash("User has been inactivated")
            return redirect(url_for('usermanagement'))
        sts = 0
        if status == 'Active':
            sts = 0
        else:
            sts = 1
        cursor.execute("UPDATE Users_man SET Name='%s',Permissions='%s',deleted = '%s', branch = '%s',lastupdated = '%s' WHERE Email='%s'" % (
            name, permissions_to_db, sts, branch,datetime.today(), email))
        conn.commit()
        flash("User details updated Successfully")
        return redirect(url_for('usermanagement'))


@app.route('/deluser', methods=['GET', 'POST'])
@login_required
def deleteuser():
    try:
        email = request.form["Email"]
        print(email)
        cursor = conn.cursor()
        #cursor.execute("update Users_man set deleted = 1 WHERE Email='%s'" % email)
        cursor.execute("delete from users_man WHERE Email='%s'" % email)
        conn.commit()
        flash("User deleted Successfully")

    except Exception as e:

        write_to_file(e)
        flash("Can't delete the user")
        conn.rollback()
    finally:
        return redirect(url_for('usermanagement'))


# """Function to change the password in usermanagement page"""
@app.route('/pasuser', methods=['GET', 'POST'])
@login_required
def passuser():
    try:
        passwa = request.form["Password1"]
        encryptpas = de.encrypt(passwa)
        print(encryptpas)
        email = request.form["Email"]
        print(email)
        print("IN password change")
        cursor = conn.cursor()
        cursor.execute("Select * from users_man where curpassword = '%s' and email = '%s' and deleted = 0"%(encryptpas, email))

        if len(list(cursor)) > 0:
            flash("Given password should be different from the previous password.")
            print("Passwords matched")

        else:
            cursor.execute("Select * from users_man where Email = '%s' and deleted =0" %email)
            if len(list(cursor)) >0:
                cursor.execute("Update Users_man SET prevPassword = Curpassword where Email = '%s' and deleted = 0"% email)
                print("password didn't match")
                cursor.execute("UPDATE Users_man SET curPassword='%s',lastupdated = '%s' WHERE Email='%s' and deleted = 0" % (encryptpas,datetime.today(), email))
                conn.commit()
                flash("Password changed Successfully")
            else:
                flash("Cannot change the password for inactive users")

#           return redirect(url_for('usermanagement'))
    except Exception as e:
        flash("Cant change the password check the database connectivity")
        write_to_file(e)
        conn.rollback()
    finally:
        return redirect(url_for('usermanagement'))


@app.route('/emailval', methods=['GET', 'POST'])
@login_required
def emailval():
    email = request.form.get("email")
    print(email)
    response = "OK"
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM Users_man where Email = '%s'" % email)
    if len(list(cursor)) > 0:
        response = "The username is already taken"
        print("username already exists")
    return response


@app.route('/checkotp', methods=['GET', 'POST'])
def checkotp():
    if request.method == "POST":
        otp = session.get("jksadfjhk")
        print("check")
        zero = request.form["0"]
        one = request.form["1"]
        two = request.form["2"]
        three = request.form["3"]

        if str(otp) == zero + one + two + three and session.get('reset_email') == request.form["emailid"]:
            print(otp)
            session['reset'] = True
            return render_template('newpass.html')
        else:
            return """<script>alert("The OTP entered is not valid");window.location.href="/"</script>"""
    return '', 202


# last stage of resetting password

@app.route('/password_reset', methods=['GET', 'POST'])
def password_reset():
    if session.get('reset') and request.method == "POST":
        email = session.get('reset_email')
        print(email)
        db1 = conn.cursor()
        db3 = conn.cursor()
        password = request.form["password"]
        encryptpas = de.encrypt(password)
        print(encryptpas)
        time = (datetime.now() - session.get('time')).total_seconds()
        print(time)
        print(session.get('time'))
        if time < 300:
            if len(password) >= 6 and len(password) < 13:  # and not str(email).__contains__("admin"):
                db2 = conn.cursor()
                db2.execute("SELECT Name from Users_man where Email='%s' and deleted = 0" % email)
                if db2 is None:
                    return "User doesn't exists. please provide a valid email"
                db2.close()
                db3.execute("Select * from users_man where curpassword = '%s' and email = '%s' and deleted = 0 " % (encryptpas, email))

                if len(list(db3)) > 0:
                    print("Passwords matched")
                    return "Given password should be different from the previous password. So, please give the new password."

                else:
                    db1.execute("Update Users_man SET prevPassword = Curpassword where Email = '%s'"% email)
                    db1.execute("UPDATE Users_man SET curPassword='%s',lastupdated='%s',temppass=0 WHERE Email='%s'" % (encryptpas, datetime.today(), email))
                    db1.commit()
                    return "OK"

            else:
                print("password strength not adequate")
                return "Not a Strong Password"

        else:
            return """<script type="text/javascript">alert("The OTP has been expired");window.location.href="/resetpass"</script>"""


"""Function to render the template"""


@app.route('/resetpass')
def resetpass():
    return render_template('resetpass.html')


"""Function to take the action resetting the password 
This function takes the email id and sends an OTP to it"""


@app.route('/reset_pass', methods=['GET', 'POST'])
def reset_pass():
    if request.method == "POST":
        email = request.form["emailid"]
        session["reset_email"] = email
        print(email)
        cursor = conn.cursor()

        cursor.execute("SELECT * FROM Users_man where Email = '%s' and deleted = 0" % email)
        if len(list(cursor)) > 0:
            otp = random.randrange(1111, 9999, 5)
            session['jksadfjhk'] = otp
            print(otp)
            db1 = conn.cursor()
            db1.execute("Select name from users_man where Email = '%s'"% email)

            for row in db1:
                name = row[0]
                print(name)
                smtpObj = smtplib.SMTP_SSL('smtp.gmail.com', port=465)
                smtpObj.login('praneeth.amudala@manomay.biz', 'manomay72020')
                msg = MIMEMultipart()
                msg['subject'] = "Reset Password"
                msg['from'] = 'praneeth.amudala@manomay.biz'
                msg['to'] = email
                msgtext1 = MIMEText(
                    '<p>Hi %s,</p>'
                    '<p>We have received a request to reset your password.</p>'
                    '<p>Please use the below OTP to set up a new password for your account.</p>'
                    '<p><strong>%s  (OTP)</strong></p>'
                    '<p>Please ignore this email, if you did not make this request.</p>'
                    '<p>Regards</p>'
                    '<p><strong>J.S.JOHNSON</strong></p>'%(name,otp), 'html')
                msg.attach(msgtext1)
                smtpObj.sendmail('praneeth.amudala@manomay.biz', email, msg.as_string())
                session["time"] = datetime.now()
            return render_template('optgenerate.html', email=email)
        else:
            print("Email is not valid")
            return """<script type="text/javascript">alert("We couldn't find the account linked with the given email id");window.location.href="/resetpass"</script> """
    else:
        return url_for('resetpass')


@app.route("/logout")
def logout():
    session.clear()
    gc.collect()
    return redirect(url_for('login'))


@app.route('/ae', methods=['GET', 'OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def ae():
    cur = conn.cursor()
    print("---------------------------------------------------------------")
    datas = request.headers.get('Authorization')
    print("-----------------",datetime.now())
    print(datas)
    datas = datas.replace('"','')
    print(datas)
    result = False
    try:
        decoded = jwt.decode(datas, str('secret'), 'utf-8')
        perm = decoded.get('permissions')
        if ("DF,DR,ARRP".find(perm) == -1):
            print("True")
        else:
            AEName = decoded.get('Name')
            print(AEName)
            cur.execute("select * from AE where AccExe = ? order by clientname,DTE",(AEName))
            row_headers=[x[0] for x in cur.description]
            records = cur.fetchall()
            json_data=[]
            for record in records:
                json_data.append(dict(zip(row_headers,record)))
        #print(records)
            result = True
            return jsonify(json_data)
    except jwt.ExpiredSignatureError:
        print("sign")
        result = False
    if result:
        return jsonify({"result": result}), 200
    else:
        return jsonify({"result": result}), 401

@app.route('/nonae', methods=['GET', 'OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def nonae():
    cur = conn.cursor()
    datas = request.headers.get('Authorization')
    datas = datas.replace('"', '')
    print("------------------------IN HO CALL  ATOKEN ISSSSSSSS-----------------------------")
    print(datas)
    result = False
    try:
        decoded = jwt.decode(datas, str('secret'), 'utf-8')
        perm = decoded.get('permissions')
        if ("DF,ARRP".find(perm) == -1):
            print("True")
        else:
            BranchName = decoded.get('Branch')
            print(BranchName)
            cur.execute("select * from NONAE where DTE between 35 and 45 and Branch = ?  order by clientname", BranchName)
            row_headers = [x[0] for x in cur.description]
            records = cur.fetchall()
            json_data=[]
            for record in records:
                json_data.append(dict(zip(row_headers,record)))
            result = True
            print("DATA -----------------")
            print(json_data)
            return jsonify(json_data)
    except jwt.ExpiredSignatureError:
        print("sign")
        result = False
    if result:
        return jsonify({"result": result}), 200
    else:
        return jsonify({"result": result}), 401

@app.route('/nonaered', methods=['GET', 'OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def nonaered():
    cur = conn.cursor()
    datas = request.headers.get('Authorization')
    datas = datas.replace('"', '')
    print(datas)
    result = False
    try:
        decoded = jwt.decode(datas, str('secret'), 'utf-8')
        perm = decoded.get('permissions')
        if ("DR,ARRP".find(perm) == -1):
            print("True")
        else:
            BranchName = decoded.get('Branch')
            print(BranchName)
            cur.execute("select * from NONAE where DTE between -30 and 30 and Branch = ? order by clientname", BranchName)
            row_headers=[x[0] for x in cur.description]
            records = cur.fetchall()
            json_data=[]
            for record in records:
                json_data.append(dict(zip(row_headers,record)))
            result = True

            return jsonify(json_data)
    except jwt.ExpiredSignatureError:
        print("sign")
        result = False
    if result:
        return jsonify({"result": result}), 200
    else:
        return jsonify({"result": result}), 401


# AE followup

@app.route("/followup", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def data():
    Data = request.get_json()
    print(Data)
    PN = Data.get('Id')
    print(PN)
    sts = Data.get('Status')
    cur = conn.cursor()
    if sts == 'Followed up':
        n = Data.get('Note')
        m = Data.get('Mode')
        d= Data.get('Date')
        print(d)
        dto = datetime.strptime(d,'%m-%d-%Y %H:%M')
        print("----------------------------date is")
        print(dto)
        fu= int(Data.get('Followups'))
        fu=fu+1
        ed = Data.get('ExpDate')
        print(ed)
        comExpDate = '01-01-2000'
        comExpDate = datetime.strptime(comExpDate,'%m-%d-%Y')
        print(comExpDate)
        print(type(comExpDate))
        print(type(ed))
        ed = datetime.strptime(ed,'%m-%d-%Y')
        UserName = Data.get('Uname')
        if ed.date() < comExpDate.date():
            print("In Expiry Date -------------------------------------")
            ed = "null"
            cur.execute("update AE set Status = ?, mode = ?,Correspondence= ?,Followups = ?,LastFollowUp = ?, username = ?, RPAStatus = NUll where PolicyNum like ?",(sts, m, n, fu, dto,UserName, PN))
            cur.commit()
        else:
            ed = ed
            cur.execute("update AE set Status = ?, mode = ?,Correspondence= ?,Followups = ?,LastFollowUp = ?,ExpectedDate= ?, username = ?, RPAStatus = null where PolicyNum like ?",(sts, m, n, fu, dto, ed, UserName, PN))
            cur.commit()
        print(ed)


    if sts == 'Completed':
        UserName = Data.get('Uname')
        tid = Data.get('TransactionId')
        d = Data.get('Date')
        dto = datetime.strptime(d,'%m-%d-%Y %H:%M')
        c = Data.get('Checked')
        cur.execute('update AE set Status = ?,ReductionTransId = ?,DateCompleted = ?,ReductionCheckBox = ?,Username = ? where PolicyNum like ?',(sts,tid,dto,c,UserName,PN))
        cur.commit()
    return jsonify(Data), 200

# NONAE FOllow up

@app.route("/Nonfollowup", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def Nonaedata():
    Data = request.get_json()
    print(Data)
    PN = Data.get('Id')
    print(PN)
    d = Data.get('Date')
    print(d)
    dto = datetime.strptime(d, '%m-%d-%Y %H:%M')
    sts = Data.get('Status')
    print(sts)
    UserName = Data.get('Uname')
    ed = Data.get('ExpDate')
    print("------------------------------Expected date is")
    print(ed)
    comExpDate = '01-01-2000'
    comExpDate = datetime.strptime(comExpDate, '%m-%d-%Y')
    ed = datetime.strptime(ed, '%m-%d-%Y')
    cur = conn.cursor()
    if sts == 'Followed up':
        n = Data.get('Note')
        print(n)
        m = Data.get('Mode')
        print(m)
        fu= int(Data.get('Followups'))
        print(fu)
        fu=fu+1
        if ed.date() < comExpDate.date():
            ed = "null"
            cur.execute("update NONAE set Status = ?, mode = ?,Correspondence= ?, Followups = ?,DateCompleted = ?, username = ? where PolicyNum like ?",(sts,m,n,fu,dto,UserName,PN))
            cur.commit()
        else:
            ed = ed
            cur.execute("update NONAE set Status = ?, mode = ?,Correspondence= ?, Followups = ?,DateCompleted = ?, username = ?, Expecteddate = ? where PolicyNum like ?",(sts, m, n, fu, dto, UserName, ed, PN))
            cur.commit()
    return jsonify(Data), 200

@app.route("/Nonaereduction", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def Nonaereddata():
    Data = request.get_json()
    print("the data is--------------")
    print(Data)
    cur = conn.cursor()
    d = Data.get('Date')
    print(d)
    dto = datetime.strptime(d, '%m-%d-%Y %H:%M')
    PN = Data.get('Id')
    sts = Data.get('Status')
    tid = Data.get('TransactionId')
    print(tid)
    rcb = Data.get('Checked')
    UserName = Data.get('Uname')
    print(UserName)
    cur.execute("update NONAE set Status = ?, ReductionTransId = ?, ReductionCheckBox = ?,DateCompleted = ?, username = ? where PolicyNum like ?",(sts,tid,rcb,dto,UserName,PN))
    cur.commit()
    return jsonify(Data), 200


@app.route("/AEEscalation", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def AEEscalation():
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))


    csr = conn.cursor()
    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws = wb.active
    ws.freeze_panes = ws['A2']
    ws.cell(1, 1).value = ('User')
    ws['A1'].font = Font(bold=True)
    ws.cell(1, 1).border = thin_border
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 2).value = ('Expired')
    ws['B1'].font = Font(bold=True)
    ws.cell(1, 2).border = thin_border
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 3).value = ('Expiring Today')
    ws['C1'].font = Font(bold=True)
    ws.cell(1, 3).border = thin_border
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 4).value = ('Not Followed up')
    ws['D1'].font = Font(bold=True)
    ws.cell(1, 4).border = thin_border
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

    csr.execute(
        "select users.name,max(p1),max(p2),max(p3) from users_man users left join (  Select name ,0 as p1,0 as p2,0 as p3 from users_man  where Permissions like '%DF,DR%'  union Select Accexe,count(policynum) as p1,0 as p2,0 as p3  from AE where DTE < 0 and  reductioncheckbox = 0  group by Accexe union Select Accexe,0 as p1,count(policynum) as p2,0  as p3  from AE where DTE = 0 and reductioncheckbox = 0  group by Accexe union  Select ACcexe,0 as p1,0 as p2,count(policynum) as p3  from AE where DTE = 15 and Followups = 0  group by Accexe)p on p.name = users.name where  users.Permissions like '%DF,DR%' group by users.name")
    rcrds = csr.fetchall()
    print(rcrds)

    row = 2
    col = 1

    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = r[2]
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[3]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width

    dt = (datetime.today()).strftime('%d%b%y')
    file_name = 'C:/Mantra/TempFolder/''AEEscalationReport' + '_' + dt + '.xlsx'
    wb.save('C:/Mantra/TempFolder/''AEEscalationReport' + '_' + dt + '.xlsx')

    return send_file(file_name)

@app.route("/ConsolidatedActivityReport", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def ConsolidatedActivityReport():


    file_name = 'C:/Mantra/Reports/ConsolidatedActivity/ConsolidatedActivityReport.xlsx'
    print(file_name)

    return send_file(file_name)


@app.route("/ExpiredPolicies", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def ExpiredPolicies():
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    wb = Workbook()
    ws = wb.active
    ws.freeze_panes = ws['A2']
    ws.cell(1, 1).value = ('Client ID')
    ws['A1'].font = Font(bold=True)
    ws.cell(1, 1).border = thin_border
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 2).value = ('Client Name')
    ws['B1'].font = Font(bold=True)
    ws.cell(1, 2).border = thin_border
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 3).value = ('Policy Number')
    ws['C1'].font = Font(bold=True)
    ws.cell(1, 3).border = thin_border
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 4).number_format = '0.00'
    ws.cell(1, 4).value = ('Premium Due')
    ws['D1'].font = Font(bold=True)
    ws.cell(1, 4).border = thin_border
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 5).value = ('Equity Date')
    ws['E1'].font = Font(bold=True)
    ws.cell(1, 5).border = thin_border
    ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
    ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 6).value = ('Email ID')
    ws['F1'].font = Font(bold=True)
    ws.cell(1, 6).border = thin_border
    ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
    ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 7).value = ('Phone Number1')
    ws['G1'].font = Font(bold=True)
    ws.cell(1, 7).border = thin_border
    ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')
    ws['G1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

    ws.cell(1, 8).value = ('Phone Number2')
    ws['H1'].font = Font(bold=True)
    ws.cell(1, 8).border = thin_border
    ws.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center')
    ws['H1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 9).value = ('Phone Number3')
    ws['I1'].font = Font(bold=True)
    ws.cell(1, 9).border = thin_border
    ws.cell(1, 9).alignment = Alignment(horizontal='center', vertical='center')
    ws['I1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 10).value = ('User')
    ws['J1'].font = Font(bold=True)
    ws.cell(1, 10).border = thin_border
    ws.cell(1, 10).alignment = Alignment(horizontal='center', vertical='center')
    ws['J1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 11).value = ('Branch')
    ws['K1'].font = Font(bold=True)
    ws.cell(1, 11).border = thin_border
    ws.cell(1, 11).alignment = Alignment(horizontal='center', vertical='center')
    ws['K1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

    row = 2

    csr = conn.cursor()
    csr.execute(
        "Select Clientid,Clientname,PolicyNum,PremDue,EquityDate,Emailid,PhnNum1,PhnNum2,PhnNum3,ACCEXE,Branch from ae  where DTE between -30 and -1  and ReductionCheckBox = 0 order by Branch")
    rcrds = csr.fetchall()
    print(rcrds)

    col = 1
    AErecords = len(rcrds)
    print(AErecords)
    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = r[2]
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).number_format = '0.00'
        ws.cell(row, col + 3).value = float(r[3])
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[4]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 5).value = r[5]
        ws.cell(row, col + 5).border = thin_border
        ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 6).value = r[6]
        ws.cell(row, col + 6).border = thin_border
        ws.cell(row, col + 6).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 7).value = r[7]
        ws.cell(row, col + 7).border = thin_border
        ws.cell(row, col + 7).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 8).value = r[8]
        ws.cell(row, col + 8).border = thin_border
        ws.cell(row, col + 8).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 9).value = r[9]
        ws.cell(row, col + 9).border = thin_border
        ws.cell(row, col + 9).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 10).value = r[10]
        ws.cell(row, col + 10).border = thin_border
        ws.cell(row, col + 10).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    # NONAE

    csr = conn.cursor()
    csr.execute(
        "Select Clientid,Clientname,PolicyNum,PremDue,EquityDate,Emailid,PhnNum1,PhnNum2,PhnNum3,Branch from nonae where  DTE between -30 and -1 and ReductionCheckBox = 0 order by branch")
    rcrds = csr.fetchall()
    print(rcrds)
    if AErecords == 0:
        row = 1
    else:
        row = row

    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = r[2]
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).number_format = '0.00'
        ws.cell(row, col + 3).value = float(r[3])
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 4).value = r[4]
        ws.cell(row, col + 4).border = thin_border
        ws.cell(row, col + 4).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 5).value = r[5]
        ws.cell(row, col + 5).border = thin_border
        ws.cell(row, col + 5).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 6).value = r[6]
        ws.cell(row, col + 6).border = thin_border
        ws.cell(row, col + 6).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 7).value = r[7]
        ws.cell(row, col + 7).border = thin_border
        ws.cell(row, col + 7).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 8).value = r[8]
        ws.cell(row, col + 8).border = thin_border
        ws.cell(row, col + 8).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 9).value = 'Non AE'
        ws.cell(row, col + 9).border = thin_border
        ws.cell(row, col + 9).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 10).value = r[9]
        ws.cell(row, col + 10).border = thin_border
        ws.cell(row, col + 0).alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width

    dt = (datetime.today()).strftime('%d%b%y')
    file_name = 'C:/Mantra/TempFolder/''ExpiredPolicies' + '_' + dt + '.xlsx'
    wb.save('C:/Mantra/TempFolder/''ExpiredPolicies' + '_' + dt + '.xlsx')

    return send_file(file_name)

@app.route("/NONAEEscalation", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def NONAEEscalation():
    print("Connected")

    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))



    wb = openpyxl.Workbook()
    ws = wb["Sheet"]
    ws = wb.active
    ws.freeze_panes = ws['A2']
    ws.cell(1, 1).value = ('Branch')
    ws['A1'].font = Font(bold=True)
    ws.cell(1, 1).border = thin_border
    ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 2).value = ('Expired')
    ws['B1'].font = Font(bold=True)
    ws.cell(1, 2).border = thin_border
    ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 3).value = ('Expiring Today')
    ws['C1'].font = Font(bold=True)
    ws.cell(1, 3).border = thin_border
    ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
    ws.cell(1, 4).value = ('Collection Call Over Due')
    ws['D1'].font = Font(bold=True)
    ws.cell(1, 4).border = thin_border
    ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")

    csr = conn.cursor()
    csr.execute(
        "select nonae.branch,max(p1),max(p2),max(p3) from nonae join (Select Branch,count(policynum) as p1,0 as p2,0 as p3  from NONAE where DTE < 0 and reductioncheckbox = 0  group by branch union Select Branch,0 as p1,count(policynum) as p2,0  as p3  from NONAE where DTE = 0 and reductioncheckbox = 0 group by branch union Select Branch,0 as p1,0 as p2,count(policynum) as p3  from NONAE where DTE between 35 and 37 and followups = 0 group by branch) p  on p.branch = nonae.Branch group by nonae.branch")
    rcrds = csr.fetchall()
    print(rcrds)

    row = 2
    col = 1

    for r in (rcrds):
        ws.cell(row, col).value = r[0]
        ws.cell(row, col).border = thin_border
        ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 1).value = r[1]
        ws.cell(row, col + 1).border = thin_border
        ws.cell(row, col + 1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 2).value = r[2]
        ws.cell(row, col + 2).border = thin_border
        ws.cell(row, col + 2).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row, col + 3).value = r[3]
        ws.cell(row, col + 3).border = thin_border
        ws.cell(row, col + 3).alignment = Alignment(horizontal='left', vertical='center')
        row += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width

    dt = (datetime.today()).strftime('%d%b%y')
    file_name = 'C:/Mantra/TempFolder/''NONAEEscalationReport' + '_' + dt + '.xlsx'
    wb.save('C:/Mantra/TempFolder/''NONAEEscalationReport' + '_' + dt + '.xlsx')

    return send_file(file_name)


@app.route("/PendingWriteoff", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def PendingWriteoff():

    dt =  ((datetime.today()) - timedelta(1)).strftime('%d%b%y')
    file_name = 'C:/Mantra/Reports/pendingwriteoff/''PendingWriteoff'+'_'+dt+'.xlsx'
    print(file_name)

    return send_file(file_name)

@app.route("/EmailFollowUps", methods=['GET','POST','OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def EmailFollowUps():

    dt = datetime.today().strftime('%d%b%y')

    file_name = 'C:/Mantra/Reports/EmailFollowUps'+dt+'.xlsx'
    print(file_name)

    return send_file(file_name)


@app.route("/InstalmentReminderFailedReport", methods=['GET', 'POST', 'OPTIONS'])
@cross_origin(origins="http://192.193.194.40:4200", headers="*")
def InstalmentReminderFailedReport():

    dt = datetime.today().strftime('%d%b%y')
    file_name = 'C:/Mantra/Reports/InstalmentReminderFailedReport'+dt+'.xlsx'
    print(file_name)

    return send_file(file_name)


def write_to_file(e):
    file = open("log.txt", "a+")
    file.writelines(str(datetime.today()) + str(e))
    file.close()


if __name__ == '__main__':
    app.secret_key = 'random string'
    app.debug = True
    app.run(host='192.193.194.40', port=5001)
