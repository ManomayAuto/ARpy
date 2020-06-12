import openpyxl
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Alignment

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

thick_border = Border(left=Side(style='thick'),
                     right=Side(style='thick'),
                     top=Side(style='thick'),
                     bottom=Side(style='thick'))


wb =openpyxl.Workbook()
ws = wb["Sheet"]
ws = wb.active
ws.freeze_panes = ws['A2']
ws.cell(1,1).value =('Client Name')
ws['A1'].font = Font(bold=True)
ws.cell(1, 1).border = thick_border
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,2).value = ('Client ID')
ws['B1'].font = Font(bold=True)
ws.cell(1, 2).border = thick_border
ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,3).value =('Policy Number')
ws['C1'].font = Font(bold=True)
ws.cell(1, 3).border = thick_border
ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,4).value =('Premium Due')
ws['D1'].font = Font(bold=True)
ws.cell(1, 4).border = thick_border
ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,5).value =('ReductionTransid')
ws['E1'].font = Font(bold=True)
ws.cell(1, 5).border = thick_border
ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,6).value =('Processed by User')
ws['F1'].font = Font(bold=True)
ws.cell(1, 6).border = thick_border
ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,7).value =('User Type')
ws['G1'].font = Font(bold=True)
ws.cell(1, 7).border = thick_border
ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')
ws['G1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,8).value =('Branch')
ws['H1'].font = Font(bold=True)
ws.cell(1, 8).border = thick_border
ws.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center')
ws['H1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
row = 2

dt = (datetime.today())
dt = dt.strftime('%Y-%m-%d')
#dt = datetime.today().strftime('%Y-%m-%d')
print(dt)
csr = conn.cursor()
csr.execute("select username,'AE',Branch,ClientName,Policynum,premdue,ReductionTransid,clientid from AE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by username",dt)
rcrds = csr.fetchall()
print(rcrds)
col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[3]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[7]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[4]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+3).number_format= '0.00'
    ws.cell(row, col+3).value = float(r[5])
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+4).value = r[6]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+5).value = r[0]
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[1]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+7).value = r[2]
    ws.cell(row, col+7).border = thin_border
    ws.cell(row, col+7).alignment = Alignment(horizontal='left', vertical='center')
    row += 1

#NONAE

csr = conn.cursor()
csr.execute("select Username,'CS',Branch,ClientName,Policynum,premdue,ReductionTransid,clientid from NONAE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by Username",dt)
rcrds = csr.fetchall()
print(rcrds)
# if AErecords == 0:
#     row = 1
# else:
#     row = row

for r in (rcrds):
    ws.cell(row, col).value = r[3]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[7]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[4]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+3).number_format= '0.00'
    ws.cell(row, col+3).value = float(r[5])
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+4).value = r[6]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+5).value = r[0]
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[1]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+7).value = r[2]
    ws.cell(row, col+7).border = thin_border
    ws.cell(row, col+7).alignment = Alignment(horizontal='left', vertical='center')
    row += 1
for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1
     ws.column_dimensions[column].width = adjusted_width
dt =  (datetime.today()).strftime('%d%b%y')
wb.save('C:/Mantra/Reports/pendingwriteoff/''PendingWriteoff'+'_'+dt+'.xlsx')


