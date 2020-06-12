
from openpyxl import *
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

t = os.path.exists('C:/Users/consultants1/Python/JSJ_Backend/Reporting/Pendingwriteoff.xlsx')
print(t)

if t:
    book=xlrd.open_workbook('C:/Users/consultants1/Python/JSJ_Backend/Reporting/PendingWriteoff.xlsx')
    sheet=book.sheet_by_index(0)
    nr= sheet.nrows
    print(nr)
    wb = load_workbook("C:/Users/consultants1/Python/JSJ_Backend/Reporting/PendingWriteoff.xlsx")
    ws = wb["Sheet"]
    row = nr + 1

else:
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1).value =('Processed by User')
    ws.cell(1,2).value =('User Type')
    ws.cell(1,3).value =('Branch')
    ws.cell(1,4).value =('Customer Name')
    ws.cell(1,5).value =('Policy Number')
    ws.cell(1,6).value =('Premium Due')
    ws.cell(1,7).value =('ReductionTransid')
    row = 2

dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
print(dt)
csr = conn.cursor()
csr.execute("select username,'AE',Branch,ClientName,Policynum,premdue,ReductionTransid from AE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by username",dt)
rcrds = csr.fetchall()
print(rcrds)
col = 1
#AErecords = len(rcrds)
#print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+5).value = r[5]
    ws.cell(row, col+6).value = r[6]
    row += 1

#NONAE

csr = conn.cursor()
csr.execute("select Username,'CS',Branch,ClientName,Policynum,premdue,ReductionTransid from NONAE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by Username",dt)
rcrds = csr.fetchall()
print(rcrds)
#if AErecords == 0:
#    row = 2
#else:
#    row = row

for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+5).value = r[5]
    ws.cell(row, col+6).value = r[6]
    row += 1

wb.save("C:/Users/consultants1/Python/JSJ_Backend/Reporting/PendingWriteoff.xlsx")


