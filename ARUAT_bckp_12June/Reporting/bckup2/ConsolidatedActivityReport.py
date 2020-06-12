from openpyxl import *
import pyodbc
import xlsxwriter
from datetime import datetime,timedelta
import os
import xlrd
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")


t = os.path.exists('C:/Mantra/Reports/ConsolidatedActivity/ConsolidatedActivityReport.xlsx')
print(t)

if t:
    book=xlrd.open_workbook('C:/Mantra/Reports/ConsolidatedActivity/ConsolidatedActivityReport.xlsx')
    sheet=book.sheet_by_index(0)
    nr= sheet.nrows
    print(nr)
    wb = load_workbook("C:/Mantra/Reports/ConsolidatedActivity/ConsolidatedActivityReport.xlsx")
    ws = wb["Sheet"]
    row = nr + 1

else:
    wb = Workbook()
    ws = wb.active
    ws.cell(1,1).value =('User')
    ws.cell(1,2).value = ('User  Type')
    ws.cell(1,3).value =('Date')
    ws.cell(1,4).value =('Follow ups Completed')
    ws.cell(1,5).value =('Reductions Completed')
    row = 2


csr = conn.cursor()
print(datetime.today())
dt=datetime.today()
print(dt)
dt1=( dt).strftime('%Y-%m-%d')
print(dt1)

csr.execute(" Select um.name,max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like '%DF,DR%' union select ae.accexe,count(lastfollowup) as f,count(DateCompleted) as r from ae where cast(lastfollowup as date) = ? or cast(DateCompleted as date) = ? group by ae.accexe) a on um.name=a.name and cast(createddate as date)<= ? group by um.name ",dt1,dt1,dt1 )
rcrds = csr.fetchall()
print(rcrds)
AErecords = len(rcrds)
print(AErecords)


aeName = 'AE'

col = 1
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col + 1).value =  aeName
    ws.cell(row, col + 2).value =  dt1
    ws.cell(row, col + 3).value =  r[1]
    ws.cell(row, col + 4).value =  r[2]
    row += 1
    print(r[0])
    print(r[1])

#NONAE DATA - Followups

csr = conn.cursor()
csr.execute(" Select um.name,'Account',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DF' union select nonae.username,count(followups) as f,0 as r from NONAE where cast(DateCompleted as date) = ? and followups=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",dt1,dt1 )
rcrds = csr.fetchall()
print(rcrds)
NONAEfrecords = len(rcrds) + AErecords
# if AErecords == 0 :
#     row = 1
# else:
#     row = row
# col = 0
for r in (rcrds):
    ws.cell(row, col).value =  r[0]
    ws.cell(row, col + 1).value =  r[1]
    ws.cell(row, col + 2).value  = dt1
    ws.cell(row, col + 3).value =  r[2]
    ws.cell(row, col + 4).value =  r[3]
    row += 1

#NONAE DATA - Reductions

csr = conn.cursor()
csr.execute("Select um.name,'CS',max(f),max(r) from users_man um join (Select u.name,0 as f,0 as r from users_man u where u.Permissions like 'DR' union select nonae.username,0 as f,count(ReductionCheckBox) as r from NONAE where cast(DateCompleted as date) =? and ReductionCheckBox=1 group by nonae.Username) a on um.name=a.name and cast(createddate as date) <= ? group by um.name",dt1,dt1 )
rcrds = csr.fetchall()
print(rcrds)
NONAERrecords = len(rcrds) + NONAEfrecords
# if NONAEfrecords == 0 :
#     row = 1
# else:
#     row = row
# col = 0
for r in (rcrds):
    ws.cell(row, col).value =  r[0]
    ws.cell(row, col + 1).value =  r[1]
    ws.cell(row, col + 2).value  = dt1
    ws.cell(row, col + 3).value =  r[2]
    ws.cell(row, col + 4).value =  r[3]
    row += 1

for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1
     ws.column_dimensions[column].width = adjusted_width
wb.save("C:/Mantra/Reports/Consolidatedactivity/ConsolidatedActivityReport.xlsx")

