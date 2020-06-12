
from openpyxl import *
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

t = os.path.exists('C:/Mantra/Reports/Pendingwriteoff/Pendingwriteoff.xlsx')
print(t)

if t:
    book=xlrd.open_workbook('C:/Mantra/Reports/PendingWriteoff/Pendingwriteoff.xlsx')
    sheet=book.sheet_by_index(0)
    nr= sheet.nrows
    print(nr)
    wb = load_workbook("C:/Mantra/Reports/PendingWriteoff/Pendingwriteoff.xlsx")
    ws = wb["Sheet"]
    row = nr + 1

else:
    wb = Workbook()
    ws = wb.active
    ws.cell(1,1).value =('Client Name')
    ws.cell(1,2).value = ('Client ID')
    ws.cell(1,3).value =('Policy Number')
    ws.cell(1,4).value =('Premium Due')
    ws.cell(1,5).value =('ReductionTransid')
    ws.cell(1,6).value =('Processed by User')
    ws.cell(1,7).value =('User Type')
    ws.cell(1,8).value =('Branch')

    row = 2

dt = (datetime.today()) - timedelta(days= 1)
dt = dt.strftime('%Y-%m-%d')
#dt = datetime.today().strftime('%Y-%m-%d')
print(dt)
csr = conn.cursor()
csr.execute("select username,'AE',Branch,ClientName,Policynum,premdue,ReductionTransid,clientid from AE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by username",dt)
rcrds = csr.fetchall()
print(rcrds)
col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[3]
    ws.cell(row, col+1).value = r[7]
    ws.cell(row, col+2).value = r[4]
    ws.cell(row, col+3).value = r[5]
    ws.cell(row, col+4).value = r[6]
    ws.cell(row, col+5).value = r[0]
    ws.cell(row, col+6).value = r[1]
    ws.cell(row, col+7).value = r[2]
    row += 1

#NONAE

csr = conn.cursor()
csr.execute("select Username,'CS',Branch,ClientName,Policynum,premdue,ReductionTransid,clientid from NONAE where cast(datecompleted as date)= ? and ReductionCheckBox = 1 order by Username",dt)
rcrds = csr.fetchall()
print(rcrds)
# if AErecords == 0:
#     row = 1
# else:
#     row = row

for r in (rcrds):
    ws.cell(row, col).value = r[3]
    ws.cell(row, col+1).value = r[7]
    ws.cell(row, col+2).value = r[4]
    ws.cell(row, col+3).value = r[5]
    ws.cell(row, col+4).value = r[6]
    ws.cell(row, col+5).value = r[0]
    ws.cell(row, col+6).value = r[1]
    ws.cell(row, col+7).value = r[2]
    row += 1
for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2
     ws.column_dimensions[column].width = adjusted_width
wb.save("C:/Mantra/Reports/PendingWriteoff/Pendingwriteoff.xlsx")


