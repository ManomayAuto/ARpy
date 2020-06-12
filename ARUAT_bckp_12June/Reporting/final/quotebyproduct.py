from datetime import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import pyodbc
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')

branch = 'Head Office'
quote_status = 'Active'

thin_border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
right=Side(border_style=BORDER_THIN, color='00000000'),
top=Side(border_style=BORDER_THIN, color='00000000'),
bottom=Side(border_style=BORDER_THIN, color='00000000')
)

cur = conn.cursor()
select_query = "select Branch, InitiatedUsername, LastName, FirstName, Product, QuoteID, QuoteIssuanceDate, TypeofCover, [AnnualNetPremium(in $)], MobileNumber, EmailID from QuoteTransaction where Branch = ? AND Quotestatus = ? AND RPAStatus IS NULL and Datediff(d,QuoteIssuanceDate,cast(getdate()+1 as date)) between 7 and 11 order by InitiatedUsername"
rows = cur.execute((select_query), (branch, quote_status))
row = rows.fetchall()

td = date.today()
t_month_day_year = td.strftime("%d%m%y")
file_name = 'C:/Mantra/Reports/QuoteReports/quotefollowup/''quotefollowup'+'_'+t_month_day_year+'.xlsx'

header = ('S.No', 'Branch', 'User', 'Client Name', 'Product', 'Quote ID', 'Quote Issuance Date', 'Type of Cover', 'Premium', 'Client Mobile Number', 'Client Email Address')
wb = Workbook()
ws = wb.active
ws.append(header)

#Freeze panes
ws.freeze_panes = ws['A2']

#Hide Gridlines
ws.sheet_view.showGridLines = False

ws['A1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['A1'].font = Font(bold = True)
ws['A1'].alignment = Alignment(horizontal='center')
ws['B1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['B1'].font = Font(bold = True)
ws['B1'].alignment = Alignment(horizontal='center')
ws['C1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['C1'].font = Font(bold = True)
ws['C1'].alignment = Alignment(horizontal='center')
ws['D1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['D1'].font = Font(bold = True)
ws['D1'].alignment = Alignment(horizontal='center')
ws['E1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['E1'].font = Font(bold = True)
ws['E1'].alignment = Alignment(horizontal='center')
ws['F1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['F1'].font = Font(bold = True)
ws['F1'].alignment = Alignment(horizontal='center')
ws['G1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['G1'].font = Font(bold = True)
ws['G1'].alignment = Alignment(horizontal='center')
ws['H1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['H1'].font = Font(bold = True)
ws['H1'].alignment = Alignment(horizontal='center')
ws['I1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['I1'].font = Font(bold = True)
ws['I1'].alignment = Alignment(horizontal='center')
ws['J1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['J1'].font = Font(bold = True)
ws['J1'].alignment = Alignment(horizontal='center')
ws['K1'].fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
ws['K1'].font = Font(bold = True)
ws['K1'].alignment = Alignment(horizontal='center')

if row:
    final_list=[]
    count = 0
    for r in row:
        client_name=""
        quote_date=''
        quote_date1=""
        count = count + 1
        r1 = list(r)
        client_name = r1[3] + ", " + r1[2]
        quote_date = datetime.strptime(r1[6], "%Y-%m-%d")
        quote_date1 = quote_date.strftime("%b %d,%Y")
        #below try is needed in case mobile number is not provided and the function int can't accept None type
        try:
            mobile_num = int(r1[9])
        except:
            print("mobile numbr is empty for Quoteid", r1[5])
            mobile_num = None
        r2 = [count, r1[0], r1[1], client_name, r1[4], r1[5],quote_date1, r1[7], float(r1[8]), mobile_num, r1[10]]
        ws.append(r2)
        curr_row = ws.max_row
        ws.cell(curr_row, 9).number_format = '$0.000'

else:
    print("no rows found for the users")

#for each column adjust the column width to the length of the maximum value in a cell
for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
         cell.border = thin_border
     adjusted_width = (max_length + 2) * 1.1
     ws.column_dimensions[column].width = adjusted_width

#center alignment of header and left alignment of other rows
skip_count = 0
for row in ws.rows:
    if skip_count == 0:
        skip_count = skip_count + 1
    else:
        for cell1 in row:
             cell1.alignment = Alignment(horizontal='left')



wb.save(file_name)
wb.close()
cur.close()




