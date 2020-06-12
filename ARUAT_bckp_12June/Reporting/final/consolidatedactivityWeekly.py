from openpyxl import *
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd
from win32com.client import Dispatch
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

wb = Workbook()
ws = wb.active
ws.freeze_panes = ws['A2']
ws.cell(1, 1).value = ('User')
ws['A1'].font = Font(bold=True)
ws.cell(1, 1).border = thin_border
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 2).value = ('Account Type')
ws['B1'].font = Font(bold=True)
ws.cell(1, 2).border = thin_border
ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 3).value = ('Date')
ws['C1'].font = Font(bold=True)
ws.cell(1, 3).border = thin_border
ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 4).value = ('Daily Activity')
ws['D1'].font = Font(bold=True)
ws.cell(1, 4).border = thin_border
ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")


row = 2

csr = conn.cursor()
csr.execute("Select Branch,Accexe,Clientid,Clientname,PolicyNum,PremDue,EquityDate,DTE from ae  where  reductioncheckbox = 0 and DTE between -30 and 30  order by Branch")
rcrds = csr.fetchall()
print(rcrds)

col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')