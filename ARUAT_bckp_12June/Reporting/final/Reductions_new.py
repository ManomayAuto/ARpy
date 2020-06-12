from openpyxl import *
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd
from win32com.client import Dispatch
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Alignment

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

wb = Workbook()
ws = wb.active
ws.freeze_panes = ws['A2']
ws.cell(1, 1).value = ('Branch')
ws['A1'].font = Font(bold=True)
ws.cell(1, 1).border = thin_border
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 2).value = ('User')
ws['B1'].font = Font(bold=True)
ws.cell(1, 2).border = thin_border
ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 3).value = ('Client ID')
ws['C1'].font = Font(bold=True)
ws.cell(1, 3).border = thin_border
ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 4).value = ('Client Name')
ws['D1'].font = Font(bold=True)
ws.cell(1, 4).border = thin_border
ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 5).value = ('Policy Number')
ws['E1'].font = Font(bold=True)
ws.cell(1, 5).border = thin_border
ws.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
ws['E1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 6).number_format = '0.00'
ws.cell(1, 6).value = ('Prem Due')
ws['F1'].font = Font(bold=True)
ws.cell(1, 6).border = thin_border
ws.cell(1, 6).alignment = Alignment(horizontal='center', vertical='center')
ws['F1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 7).value = ('Equity Date')
ws['G1'].font = Font(bold=True)
ws.cell(1, 7).border = thin_border
ws.cell(1, 7).alignment = Alignment(horizontal='center', vertical='center')
ws['G1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 8).value = ('DTE')
ws['H1'].font = Font(bold=True)
ws.cell(1, 8).border = thin_border
ws.cell(1, 8).alignment = Alignment(horizontal='center', vertical='center')
ws['H1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 9).value = ('Transaction ID')
ws['I1'].font = Font(bold=True)
ws.cell(1, 9).border = thin_border
ws.cell(1, 9).alignment = Alignment(horizontal='center', vertical='center')
ws['I1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
ws.cell(1, 10).value = ('Reduced By')
ws['J1'].font = Font(bold=True)
ws.cell(1, 10).border = thin_border
ws.cell(1, 10).alignment = Alignment(horizontal='center', vertical='center')
ws['J1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")



row = 2

csr = conn.cursor()
csr.execute("Select Branch,Accexe,Clientid,Clientname,PolicyNum,PremDue,EquityDate,DTE from ae  where  reductioncheckbox = 0 and DTE between -30 and 30  order by Branch")
rcrds = csr.fetchall()
print(rcrds)

col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+5).number_format= '0.00'
    ws.cell(row, col+5).value = float(r[5])
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[6]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    #ws.cell(row, col + 7).number_format = '0.00'
    ws.cell(row, col+7).value = int(r[7])
    ws.cell(row, col+7).border = thin_border
    ws.cell(row, col+7).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 8).value = ''
    ws.cell(row, col+8).border = thin_border
    ws.cell(row, col+8).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 9).value = ''
    ws.cell(row, col+9).border = thin_border
    ws.cell(row, col+9).alignment = Alignment(horizontal='left', vertical='center')
    row +=1



csr = conn.cursor()
csr.execute("Select Branch,ACCEXE,Clientid,Clientname,PolicyNum,PremDue,EquityDate,DTE,reductiontransid,username from ae  where  reductioncheckbox = 1 and DTE between -30 and 30   order by Branch")
rcrds = csr.fetchall()
print(rcrds)

col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+3).number_format= '0.00'
    ws.cell(row, col+5).value = float(r[5])
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[6]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    #ws.cell(row, col + 7).number_format = '0.00'
    ws.cell(row, col+7).value = int(r[7])
    ws.cell(row, col+7).border = thin_border
    ws.cell(row, col+7).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 8).value = r[8]
    ws.cell(row, col+8).border = thin_border
    ws.cell(row, col+8).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 9).value = r[9]
    ws.cell(row, col+9).border = thin_border
    ws.cell(row, col+9).alignment = Alignment(horizontal='left', vertical='center')

    row += 1

#NONAE


csr = conn.cursor()
csr.execute("Select Branch,Clientid,Clientname,PolicyNum,PremDue,EquityDate,DTE from nonae where reductioncheckbox = 0  and  DTE between -30 and 30  order by branch")
rcrds = csr.fetchall()
print(rcrds)
# if AErecords == 0:
#     row = 1
# else:
#     row = row

for r in (rcrds):

    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 1).value = 'Non AE'
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 2).value = r[1]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row, col+3).value = r[2]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 4).value = r[3]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+5).number_format= '0.00'
    ws.cell(row, col + 5).value = float(r[4])
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[5]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    #ws.cell(row, col + 7).number_format = '0'
    ws.cell(row, col + 7).value = int(r[6])
    ws.cell(row, col+7).border = thin_border
    ws.cell(row, col+7).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 8).value = ''
    ws.cell(row, col+8).border = thin_border
    ws.cell(row, col+8).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 9).value = ''
    ws.cell(row, col+9).border = thin_border
    ws.cell(row, col+9).alignment = Alignment(horizontal='left', vertical='center')
    row +=1



csr = conn.cursor()
csr.execute("Select Branch,Clientid,Clientname,PolicyNum,PremDue,EquityDate,DTE,reductiontransid,username from nonae where reductioncheckbox = 1 and   DTE between -30 and 30  order by branch")
rcrds = csr.fetchall()
print(rcrds)
# if AErecords == 0:
#     row = 1
# else:
#     row = row

for r in (rcrds):

    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 1).value = 'Non AE'
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 2).value = r[1]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')

    ws.cell(row, col+3).value = r[2]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 4).value = r[3]
    ws.cell(row, col+4).border = thin_border
    ws.cell(row, col+4).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+5).number_format= '0.00'
    ws.cell(row, col + 5).value = float(r[4])
    ws.cell(row, col+5).border = thin_border
    ws.cell(row, col+5).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+6).value = r[5]
    ws.cell(row, col+6).border = thin_border
    ws.cell(row, col+6).alignment = Alignment(horizontal='left', vertical='center')
    #ws.cell(row, col + 7).number_format = '0'
    ws.cell(row, col + 7).value = int(r[6])
    ws.cell(row, col + 7).border = thin_border
    ws.cell(row, col + 7).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 8).value = r[7]
    ws.cell(row, col+8).border = thin_border
    ws.cell(row, col+8).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col + 9).value = r[8]
    ws.cell(row, col+9).border = thin_border
    ws.cell(row, col+9).alignment = Alignment(horizontal='left', vertical='center')

    row += 1
for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1
     ws.column_dimensions[column].width = adjusted_width

dt = (datetime.today()).strftime('%d%b%y')
print(dt)
wb.save('C:/Mantra/Reports/Reductions' + '_' + dt + '.xlsx')



