import openpyxl
import pyodbc
from datetime import datetime, timedelta

from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill,Alignment

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))




csr = conn.cursor()
wb =openpyxl.Workbook()
ws = wb["Sheet"]
ws = wb.active
ws.freeze_panes = ws['A2']
ws.cell(1,1).value =('Branch')
ws['A1'].font = Font(bold=True)
ws.cell(1, 1).border = thin_border
ws.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
ws['A1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,2).value = ('Expired')
ws['B1'].font = Font(bold=True)
ws.cell(1, 2).border = thin_border
ws.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
ws['B1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,3).value =('Expiring Today')
ws['C1'].font = Font(bold=True)
ws.cell(1, 3).border = thin_border
ws.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
ws['C1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")
ws.cell(1,4).value =('Collection Call Over Due')
ws['D1'].font = Font(bold=True)
ws.cell(1, 4).border = thin_border
ws.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
ws['D1'].fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2",fill_type="solid")


csr = conn.cursor()
csr.execute("select nonae.branch,max(p1),max(p2),max(p3) from nonae join (Select Branch,count(policynum) as p1,0 as p2,0 as p3  from NONAE where DTE < 0 and reductioncheckbox = 0  group by branch union Select Branch,0 as p1,count(policynum) as p2,0  as p3  from NONAE where DTE = 0 and reductioncheckbox = 0 group by branch union Select Branch,0 as p1,0 as p2,count(policynum) as p3  from NONAE where DTE between 35 and 37 and followups = 0 group by branch) p  on p.branch = nonae.Branch group by nonae.branch")
rcrds = csr.fetchall()
print(rcrds)

row = 2
col = 1

for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col).border = thin_border
    ws.cell(row, col).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+1).border = thin_border
    ws.cell(row, col+1).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+2).border = thin_border
    ws.cell(row, col+2).alignment = Alignment(horizontal='left', vertical='center')
    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+3).border = thin_border
    ws.cell(row, col+3).alignment = Alignment(horizontal='left', vertical='center')
    row += 1

for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1
     ws.column_dimensions[column].width = adjusted_width

dt = (datetime.today()).strftime('%d%b%y')
wb.save('C:/Mantra/Reports/NONAEEscalation/''NONAEEscalationReport'+'_'+dt+'.xlsx')