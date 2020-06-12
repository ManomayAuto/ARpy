from datetime import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
import pyodbc
conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')

branch = 'Head Office'
permission = 'QI'

thin_border = Border(left=Side(border_style=BORDER_THIN, color='00000000'),
right=Side(border_style=BORDER_THIN, color='00000000'),
top=Side(border_style=BORDER_THIN, color='00000000'),
bottom=Side(border_style=BORDER_THIN, color='00000000')
)

#select all the products and display them on the header of excel file
cur = conn.cursor()
select_query = "select Description from Product"
rows_prod = cur.execute(select_query)
row_prod = rows_prod.fetchall()

header=[]
header1=['S.No', 'Branch', 'Name']
for each in range(len(row_prod)):
    element = str(row_prod[each][0])
    header1.append(element)
    header.append(element) #need this header in the next check
#print("header is", header)

#generate a file with date, month and year of current date
#write the header to the excel file
td = date.today()
t_month_year = td.strftime("%d%m%y")
file_name = 'C:/Mantra/Reports/QuoteReports/quotetracker/''quotetracker'+'_'+t_month_year+'.xlsx'
wb = Workbook()
ws = wb.active
ws.append(header1)

#Freeze panes
ws.freeze_panes = ws['A2']

#Hide Gridlines
ws.sheet_view.showGridLines = False

#auto fill the header columns with color
for each_header1 in range(1,len(header1)+1):
    ws.cell(1,each_header1).fill = PatternFill(start_color = "BFEFFF", end_color = "BFEFFF", fill_type = 'solid')
    ws.cell(1, each_header1).font = Font(bold=True)
    ws.cell(1, each_header1).alignment = Alignment(horizontal='center')

#Select all the users and also all the users having products
cur1 = conn.cursor()
select_query1= "SELECT u.Branch, u.Name, q.Product, count(q.Product) FROM Users_man as u LEFT JOIN QuoteTransaction as q ON u.Name = q.InitiatedUsername where u.Branch = ? AND u.Permissions LIKE ? and  month(q.QuoteIssuanceDate) = MONTH(CURRENT_TIMESTAMP) GROUP BY u.Branch, u.Name, q.Product union SELECT u.Branch, u.Name, q.Product, count(q.Product) as f FROM Users_man as u LEFT JOIN QuoteTransaction as q ON u.Name = q.InitiatedUsername where u.Branch = ? AND u.Permissions LIKE ? GROUP BY u.Branch, u.Name, q.Product"
rows = cur1.execute((select_query1), (branch, '%'+permission+'%',branch, '%'+permission+'%'))
row = rows.fetchall()

print("row is", row)

# merge the rows from previous query 'row' into a single row with same branch and same name
# as they belong to a single user
if row:
    mainlist2 = []
    row1 = list(row)
    rcrd_prev = list(row1[0])
    for each in range(1, len(list(row1))):
        rcrd = list(row1[each])
        if rcrd_prev[0] == rcrd[0] and rcrd_prev[1] == rcrd[1]:
            for each1 in range(2,len(rcrd)):
                rcrd_prev.append(rcrd[each1])
        else:
            mainlist2.append(rcrd_prev)
            rcrd_prev = rcrd

    #need this to print the last record
    mainlist2.append(rcrd_prev)

    #print each user and the corresponding products on the excel
    count = 0
    for each3 in range(len(mainlist2)):
        list1 = mainlist2[each3]
        count = count + 1
        sub_list = [count, list1[0], list1[1]]
        for each4 in range(len(header)):
            if header[each4] in list1:
                ind = list1.index(header[each4])
                sub_list.append(list1[ind + 1])
            else:
                sub_list.append(0)
        ws.append(sub_list)

else:
    print("no rows found for the users")

#for each column adjust the column width to the length of the maximum value in a cell
for col in ws.columns:
     max_length = 0
     column = col[0].column
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(str(cell.value))
         except:
             pass
         cell.border = thin_border
     adjusted_width = (max_length + 2) * 1.1
     ws.column_dimensions[column].width = adjusted_width

#center alignment of header and left alignment of other rows
skip_count = 0
for row in ws.rows:
    if skip_count == 0:
        skip_count = skip_count + 1
    else:
        for cell1 in row:
             cell1.alignment = Alignment(horizontal='left')


#save everything and close the file and database
wb.save(file_name)
wb.close()
cur.close()
cur1.close()