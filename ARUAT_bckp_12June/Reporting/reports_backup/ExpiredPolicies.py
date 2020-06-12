from openpyxl import *
import pyodbc
from datetime import datetime, timedelta
import os
import xlrd

conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=192.193.194.40;'
                      'Database=MantraDB;'
                      'Trusted_Connection=no;')
print("Connected")

t = os.path.exists('C:/Mantra/Reports/Expiredpolicies.xlsx')
print(t)

if t:
    book=xlrd.open_workbook('C:/Mantra/Reports/Expiredpolicies.xlsx')
    sheet=book.sheet_by_index(0)
    nr= sheet.nrows
    print(nr)
    wb = load_workbook("C:/Mantra/Reports/Expiredpolicies.xlsx")
    ws = wb["Sheet"]
    row = nr + 1
else:
    wb = Workbook()
    ws = wb.active
    ws.cell(1,1).value =('Client Name')
    ws.cell(1,2).value = ('Client ID')
    ws.cell(1,3).value =('Email Id')
    ws.cell(1,4).value =('Phone Number1')
    ws.cell(1,5).value =('Phone Number2')
    ws.cell(1,6).value =('Phone Number3')
    ws.cell(1,7).value =('Policy Number')
    ws.cell(1,8).value =('Premium Due')
    ws.cell(1,9).value =('Equity Date')
    ws.cell(1,10).value =('Branch')
    ws.cell(1,11).value =('User')

    row = 2
dt = (datetime.today() - timedelta(days=1)).strftime('%Y-%m-%d')
print(dt)
csr = conn.cursor()
csr.execute("Select Clientname,Clientid,EmailId,PhnNum1,PhnNum2,PhnNum3,PolicyNum,PremDue,EquityDate,Branch,AccEXe from ae  where cast(equitydate as date)= ? and ReductionCheckBox = 0 order by Branch",dt)
rcrds = csr.fetchall()
print(rcrds)

col = 1
AErecords = len(rcrds)
print(AErecords)
for r in (rcrds):
    ws.cell(row, col).value = r[0]
    ws.cell(row, col+1).value = r[1]
    ws.cell(row, col+2).value = r[2]
    ws.cell(row, col+3).value = r[3]
    ws.cell(row, col+4).value = r[4]
    ws.cell(row, col+5).value = r[5]
    ws.cell(row, col+6).value = r[6]
    ws.cell(row, col+7).value = r[7]
    ws.cell(row, col + 8).value = r[8]
    ws.cell(row, col + 9).value = r[9]
    ws.cell(row, col + 10).value = r[10]
    row += 1

#NONAE

csr = conn.cursor()
csr.execute("Select Clientname,clientid,EmailId,PhnNum1,PhnNum2,PhnNum3,PolicyNum,PremDue,EquityDate,Branch from nonae where cast(Equitydate as date)= ? and ReductionCheckBox = 0 order by branch",dt)
rcrds = csr.fetchall()
print(rcrds)
if AErecords == 0:
    row = 1
else:
    row = row

for r in (rcrds):

    ws.cell(row, col).value = r[0]
    ws.cell(row, col + 1).value = r[1]
    ws.cell(row, col + 2).value = r[2]
    ws.cell(row, col + 3).value = r[3]
    ws.cell(row, col + 4).value = r[4]
    ws.cell(row, col + 5).value = r[5]
    ws.cell(row, col+6).value = r[6]

    ws.cell(row, col + 7).value = r[7]
    ws.cell(row, col + 8).value = r[8]
    ws.cell(row, col + 9).value = r[9]
    ws.cell(row, col + 10).value = 'Non AE'
    row += 1

wb.save("C:/Mantra/Reports/Expiredpolicies.xlsx")


